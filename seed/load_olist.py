# -*- coding: utf-8 -*-
"""Etapa 1 — staging: CSVs Olist + planilha CMED + JSON IBGE → stg_* no Supabase."""
import json
import unicodedata

import pandas as pd
from openpyxl import load_workbook

from seed.config import DATA, connect, copy_df, run_sql_file, REPO


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return " ".join(s.upper().split())


def load_cmed() -> pd.DataFrame:
    wb = load_workbook(DATA / "cmed/cmed_pmc_20260710.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header, header_idx = None, {}
    for row in rows:
        cells = [norm(c) if c else "" for c in row]
        if "PRODUTO" in cells and any("APRESENTACAO" in c for c in cells):
            header = cells
            break
    if header is None:
        raise RuntimeError("cabecalho da CMED nao encontrado")
    for i, name in enumerate(header):
        if name.startswith("SUBSTANCIA") and "substancia" not in header_idx:
            header_idx["substancia"] = i
        elif name == "CNPJ":
            header_idx["cnpj_lab"] = i
        elif name.startswith("LABORATORIO"):
            header_idx["laboratorio"] = i
        elif name == "PRODUTO":
            header_idx["produto"] = i
        elif name.startswith("APRESENTACAO"):
            header_idx["apresentacao"] = i
    assert len(header_idx) == 5, f"colunas CMED incompletas: {header_idx}"
    recs = []
    for row in rows:  # continua de onde o header parou
        get = lambda k: (row[header_idx[k]] or "") if len(row) > header_idx[k] else ""
        produto, apres = str(get("produto")).strip(), str(get("apresentacao")).strip()
        if not produto or not apres:
            continue
        recs.append({
            "substancia": norm(get("substancia"))[:200],
            "cnpj_lab": "".join(ch for ch in str(get("cnpj_lab")) if ch.isdigit()),
            "laboratorio": norm(get("laboratorio"))[:120],
            "produto": norm(produto)[:80],
            "apresentacao": norm(apres)[:150],
        })
    wb.close()
    df = pd.DataFrame(recs).drop_duplicates(subset=["produto", "apresentacao", "cnpj_lab"])
    return df[df["cnpj_lab"].str.len() == 14].reset_index(drop=True)


def main() -> None:
    conn = connect()
    print("[ddl] criando schema (drop + create)...")
    run_sql_file(conn, REPO / "seed/ddl.sql")

    o = DATA / "olist"
    orders = pd.read_csv(o / "olist_orders_dataset.csv",
                         usecols=["order_id", "customer_id", "order_status", "order_purchase_timestamp"])
    print(f"[stg] orders: {copy_df(conn, 'stg_orders', orders):,}")

    items = pd.read_csv(o / "olist_order_items_dataset.csv",
                        usecols=["order_id", "order_item_id", "product_id", "price", "freight_value"])
    print(f"[stg] order_items: {copy_df(conn, 'stg_order_items', items):,}")

    cust = pd.read_csv(o / "olist_customers_dataset.csv",
                       usecols=["customer_id", "customer_unique_id", "customer_city", "customer_state"])
    cust["customer_city"] = cust["customer_city"].map(norm)
    print(f"[stg] customers: {copy_df(conn, 'stg_customers', cust):,}")

    prod = pd.read_csv(o / "olist_products_dataset.csv", usecols=["product_id", "product_category_name"])
    print(f"[stg] products: {copy_df(conn, 'stg_products', prod):,}")

    pay = pd.read_csv(o / "olist_order_payments_dataset.csv", usecols=["order_id", "payment_type", "payment_value"])
    pay = (pay.sort_values("payment_value", ascending=False)
              .drop_duplicates(subset=["order_id"])[["order_id", "payment_type"]])
    print(f"[stg] payments_agg (tipo dominante por pedido): {copy_df(conn, 'stg_payments_agg', pay):,}")

    cmed = load_cmed()
    print(f"[stg] cmed: {copy_df(conn, 'stg_cmed', cmed):,} apresentacoes | "
          f"{cmed['produto'].nunique():,} produtos-pai | {cmed['laboratorio'].nunique():,} laboratorios")

    munis = json.loads((DATA / "ibge/municipios.json").read_text(encoding="utf-8"))
    ibge = pd.DataFrame([{
        "codigo_ibge": m["id"],
        "nome": norm(m["nome"]),
        "uf": m["microrregiao"]["mesorregiao"]["UF"]["sigla"] if m.get("microrregiao") else None,
        "regiao": m["microrregiao"]["mesorregiao"]["UF"]["regiao"]["nome"] if m.get("microrregiao") else None,
    } for m in munis]).dropna(subset=["uf"])
    print(f"[stg] ibge: {copy_df(conn, 'stg_ibge', ibge):,}")

    conn.close()
    print("ETAPA 1 (staging) OK")


if __name__ == "__main__":
    main()
